from tkinter import filedialog
from tkinter import *
import image_processing
from Bucket import *
import globals
from openpyxl import *
import numpy
from g_code_generator import *
from PIL import ImageTk, Image

class Window(Frame):

    bucket_entry = []
    bead_labels = []
    def __init__(self, master):
        Frame.__init__(self, master)
        self.master = master
        self.init_window(master)

    #Creation of init_window
    def init_window(self, master):
        get_excel_data()
        for i in range(len(globals.bucket_array)):

            self.bucket_entry.append(LabelFrame(master, width=40,height = 40,borderwidth=1, relief="solid"))
            self.bucket_entry[i].grid(row=globals.bucket_array[i].bucket_y, column=globals.bucket_array[i].bucket_x,
                                      rowspan=1,columnspan=1, pady=2, padx=2)
            self.bucket_entry[i]["bg"] = globals.bucket_array[i].bucket_hex_color

            text = Label(self.bucket_entry[i],width=5,text = ("Helvetica", 8))
            text['text'] = globals.bucket_array[i].bucket_color_name
            text['bg'] = text.master['bg']
            text.grid(row=0, column=0)

            self.bead_labels.append(Label(self.bucket_entry[i],width=5,text = ("Helvetica", 8)))
            self.bead_labels[-1]['text'] = "0"
            self.bead_labels[-1]['bg'] = text.master['bg']
            self.bead_labels[-1].grid(row=1, column=0)

            globals.check_box_values[i] = IntVar(value=globals.bucket_array[i].bucket_available)
            c = Checkbutton(self.bucket_entry[i], variable=globals.check_box_values[i],onvalue=1, offvalue=0)
            c['bg'] = c.master['bg']
            c.grid(row = 2, column=0)

        self.get_words = Button(master)
        self.get_words["text"] = "Convert Image"
        self.get_words["command"] = self.chooseImage
        self.get_words["font"]= ("Helvetica", 14)
        self.get_words.grid(row=10,column=10,columnspan=2,pady=5)

        self.startcord = Label(master)
        self.startcord["text"] = "Panel start y:"
        self.startcord["font"]= ("Helvetica", 14)
        self.startcord.grid(row=10,column=12,columnspan=1,pady=5)

        self.startcordY = Entry(master, width=3)
        self.startcordY.grid(row=10,column=14)
        self.startcordY.insert(0, "0")

        self.size = Label(master)
        self.get_words["font"]= ("Helvetica", 14)
        self.get_words.grid(row=11,column=10,columnspan=2,pady=5)

        self.startcord = Label(master)
        self.startcord["text"] = "Panel (x,y):"
        self.startcord["font"]= ("Helvetica", 14)
        self.startcord.grid(row=11,column=12,columnspan=1,pady=5)

        self.xpanelnumber = Entry(master, width=3)
        self.xpanelnumber.grid(row=11,column=13)
        self.xpanelnumber.insert(0, "0")
        self.ypanelnumber = Entry(master, width=3)
        self.ypanelnumber.grid(row=11,column=14)
        self.ypanelnumber.insert(0, "0")

        self.var1 = IntVar()
        Checkbutton(master, text="generate all", variable=self.var1).grid(row=11,column=15, sticky=W)

        self.var2 = IntVar()
        Checkbutton(master, text="OVERRIDE:", variable=self.var2).grid(row=12,column=12, sticky=W)

        self.overidestart = Label(master)
        self.overidestart["text"] = "start:"
        self.overidestart["font"]= ("Helvetica", 14)
        self.overidestart.grid(row=12,column=13,columnspan=1,pady=5)

        self.overrideStartX = Entry(master, width=3)
        self.overrideStartX.grid(row=12,column=14)
        self.overrideStartX.insert(0, "0")
        self.overrideStartY = Entry(master, width=3)
        self.overrideStartY.grid(row=12,column=15)
        self.overrideStartY.insert(0, "0")

        self.overidestart = Label(master)
        self.overidestart["text"] = "End:"
        self.overidestart["font"]= ("Helvetica", 14)
        self.overidestart.grid(row=13,column=13,columnspan=1,pady=5)

        self.overrideEndX = Entry(master, width=3)
        self.overrideEndX.grid(row=13,column=14)
        self.overrideEndX.insert(0, "0")
        self.overrideEndY = Entry(master, width=3)
        self.overrideEndY.grid(row=13,column=15)
        self.overrideEndY.insert(0, "0")

        self.refresh = Button(master)
        self.refresh["text"] = "Refresh"
        self.refresh["command"] = self.set_start
        self.refresh["font"]= ("Helvetica", 12)
        self.refresh.grid(row=10,column=15,columnspan=1,pady=5)

        self.test = Button(master)
        self.test["text"] = "TEST"
        self.test["command"] = self.tester
        self.test["font"]= ("Helvetica", 8)
        self.test.grid(row=10,column=16,columnspan=1,pady=2)

        self.canvas = Canvas(master, width=600, height=600)
        self.canvas.grid(row=0,column=10,rowspan=10,columnspan=10)
    def tester(self):
        print("TESTING")
        for i in range(len(globals.bucket_array)):
            print(str(globals.check_box_values[i].get()))
    def update_bucket_amounts(self):
        for i in range(len(globals.bucket_array)):
            self.bead_labels[i]["text"] = str(globals.bucket_array[i].bucket_total_beads)
    def set_start(self):

        yoffset = int(self.startcordY.get())
        x_panel_number = int(self.xpanelnumber.get())
        y_panel_number = int(self.ypanelnumber.get())
        #break up image
        x_panels = round((globals.perler_image_width/globals.MAX_PANAL_SIZE)+.5)
        y_panels = round((globals.perler_image_height/globals.MAX_PANAL_SIZE)+.5)
        x_max_panel = round(globals.perler_image_width/x_panels)
        y_max_panel = round(globals.perler_image_height/y_panels)


        for i in range(0,x_panels):
            globals.perler_image.putpixel((i*x_max_panel,0),(0, 255, 0))
        for i in range(0,y_panels):
            globals.perler_image.putpixel((0,i*y_max_panel),(0, 255, 0))


        start_x = x_panel_number*x_max_panel
        if start_x > 0:
            start_x += 1
        end_x = start_x+x_max_panel
        if end_x >= globals.perler_image_width:
            end_x = globals.perler_image_width - 1


        start_y = y_panel_number*y_max_panel
        end_y = start_y+y_max_panel
        if end_y >= globals.perler_image_height:
            end_y = globals.perler_image_height - 1


        if self.var2.get() == 1:
            localOverrideStartX = int(self.overrideStartX.get())
            localOverrideStartY = int(self.overrideStartY.get())
            localOverrideEndX = int(self.overrideEndX.get())
            localOverrideEndY = int(self.overrideEndY.get())
            temp_array = image_processing.find_edges(localOverrideStartX, localOverrideStartY, localOverrideEndX, localOverrideEndY)
        else:
            temp_array = image_processing.find_edges(start_x,start_y,end_x,end_y)

        new_start_x = temp_array[0]
        new_start_y = temp_array[1]
        new_end_x = temp_array[2]
        new_end_y = temp_array[3]

        globals.perler_image.putpixel((new_start_x, new_start_y), (255, 0, 0))
        globals.perler_image.putpixel((new_start_x, new_end_y), (255, 0, 0))
        globals.perler_image.putpixel((new_end_x, new_end_y), (255, 0, 0))
        globals.perler_image.putpixel((new_end_x, new_start_y), (255, 0, 0))

        globals.perler_image.putpixel((new_start_x, new_start_y+yoffset), (255, 0, 0))
        ysize = new_end_y-new_start_y
        xsize = new_end_x - new_start_x
        self.overidestart["text"] = "x: " + str(xsize) + " y: " + str(ysize)

        if self.var2.get() == 1:
            #OVERRIDE
            print("OVERRIDE: " + str(localOverrideStartX) + "," + str(localOverrideStartY) + " " + str(localOverrideEndX) + "," + str(localOverrideEndY))
            generate_g_code("OVERRIDE",yoffset,new_start_x, new_start_y, new_end_x, new_end_y)
        else:
            if self.var1.get() == 0:
                generate_g_code("panel",yoffset,new_start_x,new_start_y,new_end_x,new_end_y)
            elif self.var1.get() == 1:
                print("GENERATE ALL")
                for i in range(0,y_panels):
                    for j in range(0,x_panels):

                        start_x = j*x_max_panel
                        end_x = start_x+x_max_panel
                        if end_x >= globals.perler_image_width:
                            end_x = globals.perler_image_width - 1
                        start_y = i*y_max_panel
                        end_y = start_y+y_max_panel
                        if end_y >= globals.perler_image_height:
                            end_y = globals.perler_image_height - 1

                        path_name = self.filename.rsplit('/', maxsplit=1)
                        generate_g_code(str(path_name[-1]) + str(j) + str(i),start_x,start_y,end_x,end_y)
        if globals.perler_image.height > globals.perler_image.width:
            multiplier = 600 / globals.perler_image.height
        else:
            multiplier = 600 / globals.perler_image.width
        globals.big_perler_image = globals.perler_image.resize((int(multiplier*globals.perler_image.width), int(multiplier*globals.perler_image.height)), Image.NONE)
        img = ImageTk.PhotoImage(globals.big_perler_image)
        self.canvas.create_image(0, 0, anchor=NW, image=img)
        self.canvas.image = img

    def chooseImage(self):
        self.filename = filedialog.askopenfilename(initialdir= globals.PERLER_PATH + "2. Perler Converter/", title="Select file",
                                                   filetypes=(("png files","*.png"),("all files", "*.*")))


        image_processing.convert_image(self.filename)

        self.set_start()
        #update bead numbers
        self.update_bucket_amounts()
        #generate gcode


def openFile():
    wb2 = load_workbook(globals.PERLER_PATH + '3. G-code Converter/bucket_chart.xlsm', data_only=True)
    return wb2

def get_excel_data():

    wb = openFile()
    sheet_ranges = wb['Sheet1']
    color_array = []
    for i in range (1,150):
        color_code = sheet_ranges.cell(column=12, row=i).value
        if color_code:
            name = sheet_ranges.cell(column=13, row=i).value
            r = int(sheet_ranges.cell(column=14, row=i).value)
            g = int(sheet_ranges.cell(column=15, row=i).value)
            b = int(sheet_ranges.cell(column=16, row=i).value)
            color_array.append([color_code,name,r,g,b])
    for col in range(8):
        for row in range(8):
            bucket_x = col
            bucket_y = row
            bucket_num = int(sheet_ranges.cell(column=col+2, row=4*row+2).value)
            bucket_color_code = int(sheet_ranges.cell(column=col+2, row=4*row+3).value)
            bucket_available = int(sheet_ranges.cell(column=col + 2, row=4 * row + 4).value)
            #find color code
            for i in range(len(color_array)):
                if color_array[i][0] == bucket_color_code:
                    globals.bucket_array.append(Bucket(bucket_x, bucket_y, bucket_num, bucket_color_code ,
                                                       color_array[i][1],color_array[i][2],color_array[i][3],
                                                       color_array[i][4],bucket_available))
                    break
    #get color codes for each bucket
    wb.close()
